#!/usr/bin/python3
# -*- coding:utf-8 -*-
#
# Copyright (c) 2019 SMHI, Swedish Meteorological and Hydrological Institute 
# License: MIT License (see LICENSE.txt or http://opensource.org/licenses/mit).

import pathlib
import openpyxl

class DwcaResources():
    """ """
    def __init__(self):
        """ """
        # Raw lists from Excel file.
        self.dwc_columns = [] 
        self.field_mapping = [] 
        self.dwc_keys = [] 
        self.dwc_dynamic_fields = [] 
        self.dwc_extra_fields = [] 
        self.metadata_mapping = [] 
        self.filter = [] 
        # Processed content.
        self.dwc_columns_dict = None 
        self.field_mapping_dict = None
        self.measurement_mapping_dict = None
        self.extra_fields_dict = None
        self.dwc_event_nodes = None
        self.dwc_occurrence_nodes = None
        self.event_node_fields = None
    
    def get_dwc_columns(self):
        """ """
        if not self.dwc_columns_dict:
            self.dwc_columns_dict = {}
            for row in self.dwc_columns:
                for key, value in row.items():
                    if value:
                        if not value[0] == '#':
                            if key not in self.dwc_columns_dict:
                                self.dwc_columns_dict[key] = []
                            #
                            self.dwc_columns_dict[key].append(value)
            
        return self.dwc_columns_dict
    
    def get_event_node_names(self):
        """ """
        self.get_event_nodes()
        
        return self.dwc_event_nodes.keys()
    
    def get_event_nodes(self):
        """ """
        if not self.dwc_event_nodes:
            self.dwc_event_nodes = {}
            for row_dict in self.dwc_keys:
                if row_dict.get('dwc_category', '') == 'event':
                    dwc_node = row_dict.get('dwc_node', '')
                    dwc_parent_event = row_dict.get('dwc_parent_event', '')
                    dwc_key_name = row_dict.get('dwc_key_name', '')
                    dwc_key_prefix = row_dict.get('dwc_key_prefix', '')
                     
                    if dwc_node:
                        if dwc_node not in self.dwc_event_nodes:
                            self.dwc_event_nodes[dwc_node] = {}
                            node_dict = self.dwc_event_nodes[dwc_node]
                            node_dict['dwc_node'] = dwc_node
                            node_dict['dwc_parent_event'] = dwc_parent_event
                            node_dict['dwc_key_name'] = dwc_key_name
                            node_dict['dwc_key_prefix'] = dwc_key_prefix
                            node_dict['key_fields'] = []
                            for key in ['key_1', 'key_2', 'key_3', 'key_4', 'key_5', 
                                        'key_6', 'key_7', 'key_8', 'key_9', 'key_10', ]:
                                value = row_dict.get(key, '')
                                if value:
                                    node_dict['key_fields'].append(value)
                            #    
                            print(node_dict)
             
        return self.dwc_event_nodes
     
    def get_occurrence_nodes(self):
        """ """
        if not self.dwc_occurrence_nodes:
            self.dwc_occurrence_nodes = {}
            for row_dict in self.dwc_keys:
                if row_dict.get('dwc_category', '') == 'occurrence':
                    dwc_node = row_dict.get('dwc_node', '')
                    # dwc_parent_event = row_dict.get('dwc_parent_event', '')
                    dwc_key_name = row_dict.get('dwc_key_name', '')
                    dwc_event_key = row_dict.get('dwc_event_key', '')
                    dwc_key_prefix = row_dict.get('dwc_key_prefix', '')
                     
                    if dwc_node:
                        if dwc_node not in self.dwc_occurrence_nodes:
                            self.dwc_occurrence_nodes[dwc_node] = {}
                            node_dict = self.dwc_occurrence_nodes[dwc_node]
                            node_dict['dwc_node'] = dwc_node
                            # node_dict['dwc_parent_event'] = dwc_parent_event
                            node_dict['dwc_key_name'] = dwc_key_name
                            node_dict['dwc_event_key'] = dwc_event_key
                            node_dict['dwc_key_prefix'] = dwc_key_prefix
                            node_dict['key_fields'] = []
                            for key in ['key_1', 'key_2', 'key_3', 'key_4', 'key_5', 
                                        'key_6', 'key_7', 'key_8', 'key_9', 'key_10', ]:
                                value = row_dict.get(key, '')
                                if value:
                                    node_dict['key_fields'].append(value)
                            #    
                            print(node_dict)
             
        return self.dwc_occurrence_nodes
    
#         if not self.dwc_occurrence_nodes:
#             self.dwc_occurrence_nodes = {}
#             for row_dict in self.occurrence_keys:
#                  
#                 dwc_event_node = row_dict.get('dwc_event_node', '')
#                 dwc_key = row_dict.get('dwc_key', '')
#                 dwc_event_key = row_dict.get('dwc_event_key', '')
#                 dwc_key_prefix = row_dict.get('dwc_key_prefix', '')
#                  
#                 if dwc_event_node:
#                     if dwc_event_node not in self.dwc_occurrence_nodes:
#                         self.dwc_occurrence_nodes[dwc_event_node] = {}
#                         node_dict = self.dwc_occurrence_nodes[dwc_event_node]
#                         node_dict['dwc_event_node'] = dwc_event_node
#                         node_dict['dwc_key'] = dwc_key
#                         node_dict['dwc_event_key'] = dwc_event_key
#                         node_dict['dwc_key_prefix'] = dwc_key_prefix
#                         node_dict['key_fields'] = []
#                         for key in ['key_1', 'key_2', 'key_3', 'key_4', 'key_5', 
#                                     'key_6', 'key_7', 'key_8', 'key_9', 'key_10', ]:
#                             value = row_dict.get(key, '')
#                             if value:
#                                 node_dict['key_fields'].append(value)
#                         #
#                         print(node_dict)
#          
#         return self.dwc_occurrence_nodes
    
    def get_event_node_fields(self, event_node):
        """ """
        if not self.event_node_fields:
            self.event_node_fields = {}
            for row_dict in self.field_mapping:
                dwc_node = row_dict.get('dwc_node', '')
                dwc_field = row_dict.get('dwc_field', )
                if dwc_field:
                    if dwc_node not in self.event_node_fields:
                        self.event_node_fields[dwc_node] = []
                    self.event_node_fields[dwc_node].append(dwc_field)
        # 
        return self.event_node_fields[event_node]
    
    def get_field_mapping(self):
        """ """
        if not self.field_mapping_dict:
            self.field_mapping_dict = {}
            for row_dict in self.field_mapping:
                if 'dwc_field' in row_dict:
                    self.field_mapping_dict[row_dict['dwc_field']] = row_dict
                    
# dwc_category
# dwc_node
# dwc_field
# dwc_measurement_type
# dwc_measurement_unit
# source_field
# source_parameter
# source_unit
        
        # 
        return self.field_mapping_dict
    
#     def get_parameter_mapping(self):
#         """ """
#         if not self.field_mapping_dict:
#             self.measurement_mapping_dict = {}
#             for row_dict in self.field_mapping:
#                 if 'dwc_measurement_type' in row_dict:
#                     [row_dict['dwc_measurement_type']
#                     self.measurement_mapping_dict[row_dict['dwc_measurement_type']] = row_dict
#                     
# # dwc_category
# # dwc_node
# # dwc_field
# # dwc_measurement_type
# # dwc_measurement_unit
# # source_field
# # source_parameter
# # source_unit
#         
#         # 
#         return self.field_mapping_dict
    
    def get_extra_fields(self):
        """ """
        if not self.extra_fields_dict:
            self.extra_fields_dict = {}
            for row_dict in self.field_mapping:
                dwc_field = row_dict.get('dwc_field', '')
                text = row_dict.get('text', '')
                if dwc_field and text:
                        self.extra_fields_dict[dwc_field] = text
        # 
        return self.extra_fields_dict
    
    def load_from_excel(self, excel_file_path):
        """ """
        excel_file_path = pathlib.Path(excel_file_path)
        workbook = None
        try:
            workbook = openpyxl.load_workbook(filename=excel_file_path)
            
            for sheet_name in workbook.sheetnames:
                print('Excel sheet: ', sheet_name)
                
                worksheet = workbook[sheet_name]
                
                if sheet_name == 'dwc_columns':
                    self._local_read_excel_sheet(worksheet, self.dwc_columns)
                if sheet_name == 'field_mapping':
                    self._local_read_excel_sheet(worksheet, self.field_mapping)
                if sheet_name == 'dwc_keys':
                    self._local_read_excel_sheet(worksheet, self.dwc_keys)
                if sheet_name == 'dwc_dynamic_fields':
                    self._local_read_excel_sheet(worksheet, self.dwc_dynamic_fields)
                if sheet_name == 'metadata_mapping':
                    self._local_read_excel_sheet(worksheet, self.metadata_mapping)
                if sheet_name == 'filter':
                    self._local_read_excel_sheet(worksheet, self.filter)
        finally:
            if workbook:
                workbook.close()
    
    def _local_read_excel_sheet(self, excel_sheet, target_list):
        """ Local function """
        header = []
        for rowindex, row in enumerate(excel_sheet.iter_rows()):
            if rowindex == 0:
                # Header.
                for cell in row:
                    value = cell.value
                    if value == None:
                        header.append('')
                    else:
                        header.append(str(value).strip())
                        
                if len(''.join(header)) == 0:
                    return # Header row is mandatory.
            else:
                # Row.
                newrow = []
                for cell in row:
                    value = cell.value
                    if value == None:
                        newrow.append('')
                    else:
                        newrow.append(str(value).strip())
                
                if len(''.join(newrow)) == 0:
                    continue # Don't add empty rows.
                
                row_dict = dict(zip(header, newrow))
                target_list.append(row_dict)
    
