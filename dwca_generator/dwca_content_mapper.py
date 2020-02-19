#!/usr/bin/python3
# -*- coding:utf-8 -*-
#
# Copyright (c) 2019 SMHI, Swedish Meteorological and Hydrological Institute 
# License: MIT License (see LICENSE.txt or http://opensource.org/licenses/mit).

import pathlib
import openpyxl

class DwcaContentMapper():
    """ """
    def __init__(self):
        """ """
        # Raw lists from Excel file.
        self.dwc_columns = [] 
        self.field_mapping = [] 
        self.dwc_keys = [] 
        self.dwc_dynamic_fields = [] 
        self.filter = [] 
        self.translate = [] 
        self.eml_metadata = [] 
        # Processed content.
        self.dwc_columns_dict = None 
        self.field_mapping_dict = None
        self.measurement_mapping_dict = None
        self.dwc_event_nodes = None
        self.dwc_occurrence_nodes = None
        self.dwc_emof_nodes = None
        self.event_node_fields = None
        self.occurrence_node_fields = None
        self.emof_node_fields = None
        self.filter_dict = None
        self.translate_from_source_dict = None
        self.translate_from_dwc_dict = None
    
    def get_translate_from_source(self, source_field, value):
        """ """
#         self.load_translate_dict()
        if self.translate_from_source_dict:
            if source_field in self.translate_from_source_dict:
                return self.translate_from_source_dict[source_field].get(value, value)
        return value
    
    def get_translate_from_dwc(self, dwc_field, value):
        """ """
#         self.load_translate_dict()
        if self.translate_from_dwc_dict:
            if dwc_field in self.translate_from_dwc_dict:
                return self.translate_from_dwc_dict[dwc_field].get(value, value)
        return value
    
    def get_translate_from_source_keys(self):
        """ """
        self.load_translate_dict()
        return self.translate_from_source_dict.keys()
    
    def get_translate_from_dwc_keys(self):
        """ """
        self.load_translate_dict()
        return self.translate_from_dwc_dict.keys()
    
    def load_translate_dict(self):
        """ """
        if not self.translate_from_source_dict:
            self.translate_from_source_dict = {}
            self.translate_from_dwc_dict = {}
            for row in self.translate:
                source_field = row.get('source_field', '')
                dwc_field = row.get('dwc_field', '')
                from_value = row.get('from_value', '')
                to_value = row.get('to_value', '')
                
                if source_field and from_value:
                    if source_field not in self.translate_from_source_dict:
                        self.translate_from_source_dict[source_field] = {}
                    self.translate_from_source_dict[source_field][from_value] = to_value
                
                if dwc_field and from_value:
                    if dwc_field not in self.translate_from_dwc_dict:
                        self.translate_from_dwc_dict[dwc_field] = {}
                    self.translate_from_dwc_dict[dwc_field][from_value] = to_value
    
    def get_filters(self):
        """ """
        if not self.filter_dict:
            self.filter_dict = {}
            for row in self.filter:
                source_field = row.get('source_field', '')
                include_value = row.get('include_value', '')
                exclude_value = row.get('exclude_value', '')
                
                if source_field and include_value:
                    if not include_value[0] == '#':
                        if source_field not in self.filter_dict:
                            self.filter_dict[source_field] = {}
                        if 'included_values' not in self.filter_dict[source_field]:
                            self.filter_dict[source_field]['included_values'] = []
                        self.filter_dict[source_field]['included_values'].append(include_value)
                
                if source_field and exclude_value:
                    if not exclude_value[0] == '#':
                        if source_field not in self.filter_dict:
                            self.filter_dict[source_field] = {}
                        if 'excluded_values' not in self.filter_dict[source_field]:
                            self.filter_dict[source_field]['excluded_values'] = []
                        self.filter_dict[source_field]['excluded_values'].append(exclude_value)
                    
        return self.filter_dict
    
    def get_dwc_columns(self):
        """ """
        if not self.dwc_columns_dict:
            self.dwc_columns_dict = {}
            for row in self.dwc_columns:
                for key, value in row.items():
                    if value:
                        if not value[0] == '#':
                            if key not in self.dwc_columns_dict:
                                self.dwc_columns_dict[key] = []
                            #
                            self.dwc_columns_dict[key].append(value)
            
        return self.dwc_columns_dict
    
    def get_event_node_names(self):
        """ """
        self.get_event_nodes_keys()
        
        return self.dwc_event_nodes.keys()
    
    def get_event_nodes_keys(self):
        """ """
        if not self.dwc_event_nodes:
            self.dwc_event_nodes = {}
            for row_dict in self.dwc_keys:
                if row_dict.get('dwc_category', '') == 'event':
                    dwc_event_node = row_dict.get('dwc_event_node', '')
                    dwc_parent_event = row_dict.get('dwc_parent_event', '')
                    dwc_key_name = row_dict.get('dwc_key_name', '')
                    dwc_event_key_name = row_dict.get('dwc_event_key_name', '')
                    dwc_key_prefix = row_dict.get('dwc_key_prefix', '')
                     
                    if dwc_event_node:
                        if dwc_event_node not in self.dwc_event_nodes:
                            self.dwc_event_nodes[dwc_event_node] = {}
                            node_dict = self.dwc_event_nodes[dwc_event_node]
                            node_dict['dwc_event_node'] = dwc_event_node
                            node_dict['dwc_parent_event'] = dwc_parent_event
                            node_dict['dwc_key_name'] = dwc_key_name
                            node_dict['dwc_event_key_name'] = dwc_event_key_name
                            node_dict['dwc_key_prefix'] = dwc_key_prefix
                            node_dict['key_fields'] = []
                            for key in ['key_1', 'key_2', 'key_3', 'key_4', 'key_5', 
                                        'key_6', 'key_7', 'key_8', 'key_9', 'key_10', ]:
                                value = row_dict.get(key, '')
                                if value:
                                    node_dict['key_fields'].append(value)
                            #    
                            print(node_dict)
             
        return self.dwc_event_nodes
     
    def get_occurrence_node_names(self):
        """ """
        self.get_occurrence_nodes_keys()
        
        return self.dwc_occurrence_nodes.keys()
    
    def get_occurrence_nodes_keys(self):
        """ """
        if not self.dwc_occurrence_nodes:
            self.dwc_occurrence_nodes = {}
            for row_dict in self.dwc_keys:
                if row_dict.get('dwc_category', '') == 'occurrence':
                    dwc_event_node = row_dict.get('dwc_event_node', '')
                    # dwc_parent_event = row_dict.get('dwc_parent_event', '')
                    dwc_key_name = row_dict.get('dwc_key_name', '')
                    dwc_event_key_name = row_dict.get('dwc_event_key_name', '')
                    dwc_key_prefix = row_dict.get('dwc_key_prefix', '')
                     
                    if dwc_event_node:
                        if dwc_event_node not in self.dwc_occurrence_nodes:
                            self.dwc_occurrence_nodes[dwc_event_node] = {}
                            node_dict = self.dwc_occurrence_nodes[dwc_event_node]
                            node_dict['dwc_event_node'] = dwc_event_node
                            # node_dict['dwc_parent_event'] = dwc_parent_event
                            node_dict['dwc_key_name'] = dwc_key_name
                            node_dict['dwc_event_key_name'] = dwc_event_key_name
                            node_dict['dwc_key_prefix'] = dwc_key_prefix
                            node_dict['key_fields'] = []
                            for key in ['key_1', 'key_2', 'key_3', 'key_4', 'key_5', 
                                        'key_6', 'key_7', 'key_8', 'key_9', 'key_10', ]:
                                value = row_dict.get(key, '')
                                if value:
                                    node_dict['key_fields'].append(value)
                            #    
                            print(node_dict)
             
        return self.dwc_occurrence_nodes
    
    def get_emof_node_names(self):
        """ """
        self.get_emof_nodes_keys()
        
        return self.dwc_emof_nodes.keys()
    
    def get_emof_nodes_keys(self):
        """ """
        if not self.dwc_emof_nodes:
            self.dwc_emof_nodes = {}
            for row_dict in self.dwc_keys:
                if row_dict.get('dwc_category', '') == 'emof':
                    dwc_event_node = row_dict.get('dwc_event_node', '')
                    # dwc_parent_event = row_dict.get('dwc_parent_event', '')
                    dwc_key_name = row_dict.get('dwc_key_name', '')
                    dwc_event_key_name = row_dict.get('dwc_event_key_name', '')
                    dwc_key_prefix = row_dict.get('dwc_key_prefix', '')
                      
                    if dwc_event_node:
                        if dwc_event_node not in self.dwc_emof_nodes:
                            self.dwc_emof_nodes[dwc_event_node] = {}
                            node_dict = self.dwc_emof_nodes[dwc_event_node]
                            node_dict['dwc_event_node'] = dwc_event_node
                            # node_dict['dwc_parent_event'] = dwc_parent_event
                            node_dict['dwc_key_name'] = dwc_key_name
                            node_dict['dwc_event_key_name'] = dwc_event_key_name
                            node_dict['dwc_key_prefix'] = dwc_key_prefix
                            node_dict['key_fields'] = []
                            for key in ['key_1', 'key_2', 'key_3', 'key_4', 'key_5', 
                                        'key_6', 'key_7', 'key_8', 'key_9', 'key_10', ]:
                                value = row_dict.get(key, '')
                                if value:
                                    node_dict['key_fields'].append(value)
                            #    
                            print(node_dict)
              
        return self.dwc_emof_nodes
    
#     def get_event_node_fields(self, event_node):
#         """ """
#         if not self.event_node_fields:
#             self.event_node_fields = {}
#             for row_dict in self.field_mapping:
#                 dwc_category = row_dict.get('dwc_category', '')
#                 if dwc_category == 'event':
#                     dwc_event_node = row_dict.get('dwc_event_node', '')
#                     dwc_field = row_dict.get('dwc_field', )
#                     if dwc_field:
#                         if dwc_event_node not in self.event_node_fields:
#                             self.event_node_fields[dwc_event_node] = []
#                         self.event_node_fields[dwc_event_node].append(dwc_field)
#         # 
#         return self.event_node_fields[event_node]
#     
#     def get_occurrence_node_fields(self, event_node):
#         """ """
#         if not self.occurrence_node_fields:
#             self.occurrence_node_fields = {}
#             for row_dict in self.field_mapping:
#                 dwc_category = row_dict.get('dwc_category', '')
#                 if dwc_category == 'occurrence':
#                     dwc_event_node = row_dict.get('dwc_event_node', '')
#                     dwc_field = row_dict.get('dwc_field', )
#                     if dwc_field:
#                         if dwc_event_node not in self.occurrence_node_fields:
#                             self.occurrence_node_fields[dwc_event_node] = []
#                         self.occurrence_node_fields[dwc_event_node].append(dwc_field)
#         # 
#         return self.occurrence_node_fields[event_node]
#     
#     def get_emof_node_fields(self, event_node):
#         """ """
#         if not self.emof_node_fields:
#             self.emof_node_fields = {}
#             for row_dict in self.field_mapping:
#                 dwc_category = row_dict.get('dwc_category', '')
#                 if dwc_category == 'emof':
#                     dwc_event_node = row_dict.get('dwc_event_node', '')
#                     dwc_field = row_dict.get('dwc_field', )
#                     if dwc_field:
#                         if dwc_event_node not in self.emof_node_fields:
#                             self.emof_node_fields[dwc_event_node] = []
#                         self.emof_node_fields[dwc_event_node].append(dwc_field)
#         # 
#         return self.emof_node_fields[event_node]
    
    def load_from_excel(self, excel_file_path):
        """ """
        excel_file_path = pathlib.Path(excel_file_path)
        workbook = None
        try:
            workbook = openpyxl.load_workbook(filename=excel_file_path)
            
            for sheet_name in workbook.sheetnames:
                print('Excel sheet: ', sheet_name)
                
                worksheet = workbook[sheet_name]
                
                if sheet_name == 'dwc_columns':
                    self._local_read_excel_sheet(worksheet, self.dwc_columns)
                if sheet_name == 'field_mapping':
                    self._local_read_excel_sheet(worksheet, self.field_mapping)
                if sheet_name == 'dwc_keys':
                    self._local_read_excel_sheet(worksheet, self.dwc_keys)
                if sheet_name == 'dwc_dynamic_fields':
                    self._local_read_excel_sheet(worksheet, self.dwc_dynamic_fields)
                if sheet_name == 'translate':
                    self._local_read_excel_sheet(worksheet, self.translate)
                if sheet_name == 'eml_metadata':
                    self._local_read_excel_sheet(worksheet, self.eml_metadata)
                if sheet_name == 'filter':
                    self._local_read_excel_sheet(worksheet, self.filter)
        finally:
            if workbook:
                workbook.close()
    
    def _local_read_excel_sheet(self, excel_sheet, target_list):
        """ Local function. """
        header = []
        for rowindex, row in enumerate(excel_sheet.iter_rows()):
            if rowindex == 0:
                # Header.
                for cell in row:
                    value = cell.value
                    if value == None:
                        header.append('')
                    else:
                        header.append(str(value).strip())
                        
                if len(''.join(header)) == 0:
                    return # Header row is mandatory.
            else:
                # Row.
                newrow = []
                for cell in row:
                    value = cell.value
                    if value == None:
                        newrow.append('')
                    else:
                        newrow.append(str(value).strip())
                
                if len(''.join(newrow)) == 0:
                    continue # Don't add empty rows.
                
                row_dict = dict(zip(header, newrow))
                target_list.append(row_dict)

